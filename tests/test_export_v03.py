from openpyxl import load_workbook
from sqlmodel import Session, SQLModel, create_engine

from jobtrail.models import Application, Status
from jobtrail.services.export import export_latex, export_xlsx, latex_escape


def test_latex_escape_and_content() -> None:
    engine = create_engine("sqlite://")
    SQLModel.metadata.create_all(engine)
    with Session(engine) as db:
        db.add(Application(company="A&B", role="ML_Engineer", status=Status.pending))
        db.commit()
        text = export_latex(db, status=Status.pending)
    assert latex_escape("A&B_1") == r"A\&B\_1"
    assert r"A\&B" in text
    assert r"ML\_Engineer" in text


def test_xlsx_export_file_creation(tmp_path) -> None:
    engine = create_engine("sqlite://")
    SQLModel.metadata.create_all(engine)
    with Session(engine) as db:
        db.add(Application(company="A", role="R", status=Status.pending))
        db.commit()
        path = export_xlsx(db, tmp_path / "out.xlsx")
    wb = load_workbook(path)
    assert "Applications" in wb.sheetnames
    assert "Stats" in wb.sheetnames
    assert "Followups" in wb.sheetnames
