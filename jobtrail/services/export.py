from __future__ import annotations

import csv
import io
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from sqlmodel import Session, select

from jobtrail.models import Application, Status
from jobtrail.services.followups import days_since_update, followup_candidates
from jobtrail.services.stats import stats


FIELDS = [
    "id",
    "company",
    "role",
    "location",
    "status",
    "application_date",
    "last_update_date",
    "days_since_update",
    "source",
    "job_url",
    "confidence",
    "manually_verified",
    "archived",
    "notes",
]


def app_rows(db: Session, status: Status | None = None) -> list[dict[str, object]]:
    apps = db.exec(select(Application).order_by(Application.id)).all()
    if status:
        apps = [app for app in apps if app.status == status]
    return [
        {
            "id": app.id,
            "company": app.company,
            "role": app.role,
            "location": app.location,
            "status": app.status.value,
            "application_date": app.application_date,
            "last_update_date": app.last_update_date,
            "days_since_update": days_since_update(app),
            "source": app.source,
            "job_url": app.job_url,
            "confidence": app.confidence,
            "manually_verified": app.manually_verified,
            "archived": app.archived,
            "notes": app.notes,
        }
        for app in apps
    ]


def export_csv(db: Session) -> str:
    out = io.StringIO()
    writer = csv.DictWriter(out, fieldnames=FIELDS)
    writer.writeheader()
    for row in app_rows(db):
        writer.writerow(row)
    return out.getvalue()


def export_markdown(db: Session, status: Status | None = None) -> str:
    rows = ["| ID | Company | Role | Status |", "|---:|---|---|---|"]
    for row in app_rows(db, status=status):
        rows.append(f"| {row['id']} | {row['company']} | {row['role']} | {row['status']} |")
    return "\n".join(rows) + "\n"


def latex_escape(value: object) -> str:
    text = "" if value is None else str(value)
    return (
        text.replace("\\", r"\textbackslash{}")
        .replace("&", r"\&")
        .replace("%", r"\%")
        .replace("$", r"\$")
        .replace("#", r"\#")
        .replace("_", r"\_")
        .replace("{", r"\{")
        .replace("}", r"\}")
        .replace("~", r"\textasciitilde{}")
        .replace("^", r"\textasciicircum{}")
    )


def export_latex(db: Session, status: Status | None = None) -> str:
    nl = chr(92) * 2
    lines = [f"{chr(92)}begin{{longtable}}{{rlll}}", f"ID & Company & Role & Status {nl}", f"{chr(92)}hline"]
    for row in app_rows(db, status=status):
        values = [row["id"], row["company"], row["role"], row["status"]]
        lines.append(" & ".join(latex_escape(value) for value in values) + f" {nl}")
    lines.append(f"{chr(92)}end{{longtable}}")
    return "\n".join(lines) + "\n"


def export_xlsx(db: Session, path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    ws.append(FIELDS)
    for row in app_rows(db):
        ws.append([row[field] for field in FIELDS])
    ws.freeze_panes = "A2"
    for column in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2
        ws.column_dimensions[column[0].column_letter].width = min(width, 50)

    stats_ws = wb.create_sheet("Stats")
    for key, value in stats(db).items():
        stats_ws.append([key, value])
    stats_ws.freeze_panes = "A2"

    followup_ws = wb.create_sheet("Followups")
    followup_ws.append(["id", "company", "role", "status", "days_stale", "suggested_action"])
    for item in followup_candidates(db):
        app = item.application
        followup_ws.append([app.id, app.company, app.role, app.status.value, item.days_stale, item.suggested_action])
    followup_ws.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def export_dir(base: Path) -> Path:
    path = base / "exports"
    path.mkdir(parents=True, exist_ok=True)
    return path


def timestamped_name(suffix: str) -> str:
    return f"jobtrail-{datetime.now(UTC):%Y%m%d-%H%M%S}.{suffix}"
